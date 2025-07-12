from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Q # Import Q for complex lookups
from .models import (
    LandRegistration,
    SurveyPayment,
    LandSurvey,
    TaxPayment,
    LandMapping,
    Approval,
    PasswordResetRequest,
    Notification
)
from .forms import (
    LandRegistrationForm,
    GiftLandRegistrationForm,
    InheritanceLandRegistrationForm,
    SurveyPaymentForm,
    LandSurveyForm,
    TaxPaymentForm,
    LandMappingForm,
    ApprovalForm,
    DirectorApprovalForm,
    SecretaryApprovalForm,
    DeputyMayorApprovalForm,
    MayorApprovalForm,
    ProfileForm,
    UserEditForm,
    UserCreateForm,
    UsernamePasswordResetForm,
    SetPasswordForm
)
from django.urls import reverse
from django.utils import timezone
import uuid
from datetime import datetime, timedelta
from django.db.models import Count, Sum, Avg
import csv
from django.contrib.auth.models import User
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.models import Group

from django.contrib.auth.tokens import default_token_generator
from django.core.mail import send_mail
from django.conf import settings

from django.contrib.auth.views import PasswordResetConfirmView
from django.http import JsonResponse

from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table as RLTable, TableStyle, SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO

def check_pending_password_reset(request):
    username = request.GET.get('username')
    if not username:
        return JsonResponse({'status': 'error', 'message': 'No username provided.'})
    try:
        user = User.objects.get(username=username)
        approved_request = PasswordResetRequest.objects.filter(user=user, status='approved').first()
        if approved_request:
            return JsonResponse({
                'status': 'approved',
                'requested_at': approved_request.requested_at.strftime('%B %d, %Y at %I:%M %p'),
                'message': f'Your password reset request was approved on {approved_request.requested_at.strftime("%B %d, %Y at %I:%M %p")}. You can now set a new password.'
            })
        pending_request = PasswordResetRequest.get_pending_request(user)
        if pending_request:
            return JsonResponse({
                'status': 'pending',
                'requested_at': pending_request.requested_at.strftime('%B %d, %Y at %I:%M %p'),
                'message': f'You have a password reset request submitted on {pending_request.requested_at.strftime("%B %d, %Y at %I:%M %p")} that is currently pending approval. Please wait for administrator review before submitting a new request.'
            })
        else:
            return JsonResponse({'status': 'ok'})
    except User.DoesNotExist:
        return JsonResponse({'status': 'not_found', 'message': 'User does not exist.'})

@login_required
def dashboard(request):
    registrations = LandRegistration.objects.filter(user=request.user)
    approved_count = registrations.filter(status='approved').count()
    pending_count = registrations.filter(status='pending').count()
    rejected_count = registrations.filter(status='rejected').count()
    completed_count = registrations.filter(status='completed').count()
    
    # Get notifications for superusers
    notifications_list = []
    if request.user.is_superuser:
        notifications_list = Notification.objects.filter(user=request.user, is_read=False)[:5]

    # Filtering
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status = request.GET.get('status')
    region = request.GET.get('region')
    if start_date:
        registrations = registrations.filter(register_date__gte=start_date)
    if end_date:
        registrations = registrations.filter(register_date__lte=end_date)
    if status and status != 'all':
        registrations = registrations.filter(status=status)
    if region and region != 'all':
        registrations = registrations.filter(land_region__icontains=region)

    # Sorting
    sort_by = request.GET.get('sort_by', 'register_date')
    sort_dir = request.GET.get('sort_dir', 'desc')
    if sort_by not in ['register_date', 'status', 'buyer_full_name', 'land_code', 'land_region']:
        sort_by = 'register_date'
    order = '-' + sort_by if sort_dir == 'desc' else sort_by
    registrations = registrations.order_by(order)

    # Payment stats
    survey_payments = SurveyPayment.objects.filter(land_registration__user=request.user)
    tax_payments = TaxPayment.objects.filter(land_registration__user=request.user)
    paid_survey_count = survey_payments.filter(payment_status='paid').count()
    paid_tax_count = tax_payments.filter(payment_status='paid').count()

    # Recent activity (last 5 registrations, filtered and sorted)
    recent_registrations = registrations[:5]

    # Chart data: Registrations per month (last 6 months)
    today = datetime.today()
    months = [(today - timedelta(days=30*i)).strftime('%b %Y') for i in range(5, -1, -1)]
    reg_per_month = []
    for m in months:
        month, year = m.split()
        count = registrations.filter(register_date__month=datetime.strptime(month, '%b').month, register_date__year=int(year)).count()
        reg_per_month.append(count)

    # Approval status breakdown
    approval_statuses = ['pending', 'approved', 'rejected', 'completed']
    approval_status_counts = [registrations.filter(status=s).count() for s in approval_statuses]

    # Notifications: pending, returned, or items needing attention
    notifications = []
    if pending_count > 0:
        notifications.append(f"You have {pending_count} pending registrations.")
    returned_regs = registrations.filter(current_step__icontains='returned')
    if returned_regs.exists():
        notifications.append(f"{returned_regs.count()} registrations were returned for correction.")
    if survey_payments.filter(payment_status='pending').exists():
        notifications.append("Some survey payments are still pending.")
    if tax_payments.filter(payment_status='pending').exists():
        notifications.append("Some tax payments are still pending.")

    # For filter dropdowns
    all_regions = LandRegistration.objects.values_list('land_region', flat=True).distinct()

    context = {
        'registrations': registrations,
        'approved_count': approved_count,
        'pending_count': pending_count,
        'rejected_count': rejected_count,
        'completed_count': completed_count,
        'paid_survey_count': paid_survey_count,
        'paid_tax_count': paid_tax_count,
        'recent_registrations': recent_registrations,
        'reg_per_month': reg_per_month,
        'months': months,
        'approval_status_counts': approval_status_counts,
        'approval_statuses': approval_statuses,
        'notifications': notifications,
        'notifications_list': notifications_list,  # New notifications for superusers
        'user': request.user,
        'all_regions': all_regions,
        'selected_status': status or 'all',
        'selected_region': region or 'all',
        'start_date': start_date or '',
        'end_date': end_date or '',
        'sort_by': sort_by,
        'sort_dir': sort_dir,
    }
    return render(request, 'land_management/general/dashboard.html', context)

@login_required
def land_registration(request, registration_id=None):
    registration = None
    if registration_id:
        registration = get_object_or_404(LandRegistration, id=registration_id, user=request.user)

    if request.method == 'POST':
        try:
            if registration: # Editing an existing registration
                form = LandRegistrationForm(request.POST, request.FILES, instance=registration)
            else: # Creating a new registration
                form = LandRegistrationForm(request.POST, request.FILES)
            
            if form.is_valid():
                registration = form.save(commit=False)
                registration.user = request.user
                registration.status = 'pending'
                registration.current_step = 'registration' # Reset to initial step if returned for correction

                # Generate unique transaction_reference only for new registrations
                if not registration_id:
                    registration.transaction_reference = f"TR-{timezone.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"

                registration.save()

                # If this was a returned registration, clear approval return info
                if registration_id and hasattr(registration, 'approval') and registration.approval.return_step:
                    approval_instance = registration.approval
                    approval_instance.return_step = None
                    approval_instance.rejection_reason = None
                    approval_instance.rejection_date = None
                    approval_instance.returned_by = None
                    approval_instance.save()

                messages.success(request, f'Land registration submitted successfully! Transaction Reference: {registration.transaction_reference}')
                return redirect('land_management:survey_payment', registration_id=registration.id)
            else:
                messages.error(request, 'Please correct the errors in the form.')
        except Exception as e:
            messages.error(request, f'Error saving registration: {str(e)}')
    else: # GET request
        if registration: # Pre-fill form for existing registration
            form = LandRegistrationForm(instance=registration)
        else: # Blank form for new registration
            form = LandRegistrationForm()

    return render(request, 'land_management/registrations/land_registration.html', {'form': form, 'registration': registration})

@login_required
def survey_payment(request, registration_id):
    registration = get_object_or_404(LandRegistration, id=registration_id, user=request.user)
    
    try:
        existing_payment = SurveyPayment.objects.get(land_registration=registration)
    except SurveyPayment.DoesNotExist:
        existing_payment = None

    if request.method == 'POST':
        form = SurveyPaymentForm(request.POST, request.FILES, instance=existing_payment)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.land_registration = registration
            payment.payment_status = 'paid'
            payment.save()
            registration.current_step = 'survey_payment'
            registration.save()
            messages.success(request, 'Survey payment submitted successfully.')
            return redirect('land_management:land_survey', registration_id=registration.id)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        # Pre-fill some fields with registration data
        initial_data = {
            'payer_name': registration.buyer_full_name,
            'admin_name': request.user.get_full_name() or request.user.username,
        }
        form = SurveyPaymentForm(instance=existing_payment, initial=initial_data)

    return render(request, 'land_management/payments/survey_payment.html', {
        'form': form,
        'registration': registration
    })

@login_required
def land_survey(request, registration_id):
    registration = get_object_or_404(LandRegistration, id=registration_id, user=request.user)
    
    try:
        existing_survey = LandSurvey.objects.get(land_registration=registration)
    except LandSurvey.DoesNotExist:
        existing_survey = None

    if request.method == 'POST':
        form = LandSurveyForm(request.POST, request.FILES, instance=existing_survey)
        if form.is_valid():
            survey = form.save(commit=False)
            survey.land_registration = registration
            survey.save()
            registration.current_step = 'land_survey'
            registration.save()
            messages.success(request, 'Land survey submitted successfully.')
            return redirect('land_management:tax_payment', registration_id=registration.id)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        # Pre-fill some fields with registration data
        initial_data = {
            'land_code': registration.land_code,
            'owner_name': registration.buyer_full_name,
            'surveyor_name': request.user.get_full_name() or request.user.username,
            'survey_location': registration.land_region,
        }
        form = LandSurveyForm(instance=existing_survey, initial=initial_data)

    return render(request, 'land_management/land_details/land_survey.html', {
        'form': form,
        'registration': registration
    })

@login_required
def tax_payment(request, registration_id):
    registration = get_object_or_404(LandRegistration, id=registration_id, user=request.user)
    
    try:
        existing_payment = TaxPayment.objects.get(land_registration=registration)
    except TaxPayment.DoesNotExist:
        existing_payment = None

    if request.method == 'POST':
        form = TaxPaymentForm(request.POST, instance=existing_payment)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.land_registration = registration
            payment.payment_status = 'paid'
            payment.save()
            registration.current_step = 'tax_payment'
            registration.save()
            messages.success(request, f'Tax payment submitted successfully! You can now proceed to <a href="{reverse('land_management:land_mapping', args=[registration.id])}" class="alert-link">Land Mapping</a>.')
            # Stay on the same page, re-render with messages
            return render(request, 'land_management/payments/tax_payment.html', {
                'form': form,
                'registration': registration,
            })
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        # Pre-fill some fields with registration data
        initial_data = {
            'admin_fullname': request.user.get_full_name() or request.user.username,
            'land_owner_name': registration.buyer_full_name,
            'land_reference_no': registration.transaction_reference,
            'land_price': registration.sale_price, # Assuming land price is sale price
            'tax_price': float(registration.sale_price) * 0.05, # Example: 5% of sale price as tax
        }
        form = TaxPaymentForm(instance=existing_payment, initial=initial_data)

    return render(request, 'land_management/payments/tax_payment.html', {
        'form': form,
        'registration': registration
    })

@login_required
def land_mapping(request, registration_id):
    registration = get_object_or_404(LandRegistration, id=registration_id, user=request.user)
    
    try:
        existing_mapping = LandMapping.objects.get(land_registration=registration)
    except LandMapping.DoesNotExist:
        existing_mapping = None

    if request.method == 'POST':
        form = LandMappingForm(request.POST, request.FILES, instance=existing_mapping)
        if form.is_valid():
            mapping = form.save(commit=False)
            mapping.land_registration = registration
            mapping.save()
            registration.current_step = 'land_mapping'
            registration.save()
            messages.success(request, f'Land mapping submitted successfully! You can now proceed to <a href="{reverse('land_management:approval_process', args=[registration.id])}" class="alert-link">Approval Process</a>.')
            # Stay on the same page, re-render with messages
            return render(request, 'land_management/land_details/land_mapping.html', {
                'form': form,
                'registration': registration,
            })
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        # Pre-fill some fields with registration data
        initial_data = {
            'land_reference': registration.transaction_reference,
            'mapped_by': request.user.get_full_name() or request.user.username,
        }
        form = LandMappingForm(instance=existing_mapping, initial=initial_data)

    return render(request, 'land_management/land_details/land_mapping.html', {
        'form': form,
        'registration': registration
    })

@login_required
def approval_process(request, registration_id):
    registration = get_object_or_404(LandRegistration, id=registration_id)
    
    try:
        approval_instance = Approval.objects.get(land_registration=registration)
    except Approval.DoesNotExist:
        approval_instance = Approval.objects.create(land_registration=registration)

    # Validation: Check for all required steps and fields before proceeding with approval
    required_fields = ['buyer_full_name', 'seller_full_name', 'land_code']
    missing_fields = [field for field in required_fields if not getattr(registration, field, None)]
    if missing_fields:
        messages.error(request, f"Missing required registration fields: {', '.join(missing_fields)}. Please complete the registration.")
        return redirect('land_management:land_registration', registration_id=registration.id)

    survey_payment = getattr(registration, 'surveypayment', None)
    land_survey = getattr(registration, 'landsurvey', None)
    tax_payment = getattr(registration, 'taxpayment', None)
    land_mapping = getattr(registration, 'landmapping', None)

    if not survey_payment:
        messages.error(request, "Survey payment is missing. Please complete the survey payment before proceeding with approval.")
        return redirect('land_management:survey_payment', registration_id=registration.id)

    if not land_survey:
        messages.error(request, "Land survey is missing. Please complete the land survey before proceeding with approval.")
        return redirect('land_management:land_survey', registration_id=registration.id)

    if not tax_payment:
        messages.error(request, "Tax payment is missing. Please complete the tax payment before proceeding with approval.")
        return redirect('land_management:tax_payment', registration_id=registration.id)

    if not land_mapping:
        messages.error(request, "Land mapping is missing. Please complete the land mapping before proceeding with approval.")
        return redirect('land_management:land_mapping', registration_id=registration.id)

    current_level = approval_instance.current_approval_level
    form = None
    template_name = 'land_management/approvals/approval_process.html'

    # Fetch all related data for display
    survey_payment = getattr(registration, 'surveypayment', None)
    land_survey = getattr(registration, 'landsurvey', None)
    tax_payment = getattr(registration, 'taxpayment', None)
    land_mapping = getattr(registration, 'landmapping', None)

    if request.method == 'POST':
        if current_level == 'director_pending':
            form = DirectorApprovalForm(request.POST, request.FILES, instance=approval_instance)
        elif current_level == 'secretary_pending' and approval_instance.director_status == 'approved':
            form = SecretaryApprovalForm(request.POST, request.FILES, instance=approval_instance)
        elif current_level == 'deputy_mayor_pending' and approval_instance.secretary_status == 'approved':
            form = DeputyMayorApprovalForm(request.POST, request.FILES, instance=approval_instance)
        elif current_level == 'mayor_pending' and approval_instance.deputy_mayor_status == 'approved':
            form = MayorApprovalForm(request.POST, request.FILES, instance=approval_instance)

        if form and form.is_valid():
            form.save()
            
            # Handle returned status
            if form.cleaned_data.get('director_status') == 'returned' or \
               form.cleaned_data.get('secretary_status') == 'returned' or \
               form.cleaned_data.get('deputy_mayor_status') == 'returned' or \
               form.cleaned_data.get('mayor_status') == 'returned':
                
                # Set return information
                approval_instance.return_step = form.cleaned_data.get('return_step')
                approval_instance.rejection_reason = form.cleaned_data.get('rejection_reason')
                approval_instance.rejection_date = timezone.now()
                approval_instance.returned_by = request.user.get_full_name() or request.user.username
                approval_instance.save()

                # Update registration status
                registration.status = 'pending'
                registration.current_step = form.cleaned_data.get('return_step')
                registration.save()

                messages.warning(request, f'Registration has been returned for correction at the {approval_instance.return_step} step.')
                
                # Redirect to the appropriate step
                if approval_instance.return_step == 'registration':
                    return redirect('land_management:land_registration')
                elif approval_instance.return_step == 'survey_payment':
                    return redirect('land_management:survey_payment', registration_id=registration.id)
                elif approval_instance.return_step == 'land_survey':
                    return redirect('land_management:land_survey', registration_id=registration.id)
                elif approval_instance.return_step == 'tax_payment':
                    return redirect('land_management:tax_payment', registration_id=registration.id)
                elif approval_instance.return_step == 'land_mapping':
                    return redirect('land_management:land_mapping', registration_id=registration.id)

            # Handle normal approval flow
            if current_level == 'director_pending':
                if approval_instance.director_status == 'approved':
                    approval_instance.director_approval_date = timezone.now()
                    approval_instance.current_approval_level = 'secretary_pending'
                    messages.success(request, 'Director approval submitted. Proceed to Secretary Approval.')
                elif approval_instance.director_status == 'rejected':
                    approval_instance.current_approval_level = 'rejected'
                    registration.status = 'rejected'
                    messages.error(request, 'Director rejected the registration.')
                approval_instance.save()
                registration.save()
                return render(request, template_name, {
                    'form': form,
                    'registration': registration,
                    'approval': approval_instance,
                    'current_level': current_level,
                    'survey_payment': survey_payment,
                    'land_survey': land_survey,
                    'tax_payment': tax_payment,
                    'land_mapping': land_mapping,
                })

            elif current_level == 'secretary_pending' and approval_instance.director_status == 'approved':
                if approval_instance.secretary_status == 'approved':
                    approval_instance.secretary_approval_date = timezone.now()
                    approval_instance.current_approval_level = 'deputy_mayor_pending'
                    messages.success(request, 'Secretary approval submitted. Proceed to Deputy Mayor Approval.')
                elif approval_instance.secretary_status == 'rejected':
                    approval_instance.current_approval_level = 'rejected'
                    registration.status = 'rejected'
                    messages.error(request, 'Secretary rejected the registration.')
                approval_instance.save()
                registration.save()
                return render(request, template_name, {
                    'form': form,
                    'registration': registration,
                    'approval': approval_instance,
                    'current_level': current_level,
                    'survey_payment': survey_payment,
                    'land_survey': land_survey,
                    'tax_payment': tax_payment,
                    'land_mapping': land_mapping,
                })

            elif current_level == 'deputy_mayor_pending' and approval_instance.secretary_status == 'approved':
                if approval_instance.deputy_mayor_status == 'approved':
                    approval_instance.deputy_mayor_approval_date = timezone.now()
                    approval_instance.current_approval_level = 'mayor_pending'
                    messages.success(request, 'Deputy Mayor approval submitted. Proceed to Mayor Approval.')
                elif approval_instance.deputy_mayor_status == 'rejected':
                    approval_instance.current_approval_level = 'rejected'
                    registration.status = 'rejected'
                    messages.error(request, 'Deputy Mayor rejected the registration.')
                approval_instance.save()
                registration.save()
                return render(request, template_name, {
                    'form': form,
                    'registration': registration,
                    'approval': approval_instance,
                    'current_level': current_level,
                    'survey_payment': survey_payment,
                    'land_survey': land_survey,
                    'tax_payment': tax_payment,
                    'land_mapping': land_mapping,
                })

            elif current_level == 'mayor_pending' and approval_instance.deputy_mayor_status == 'approved':
                if approval_instance.mayor_status == 'approved':
                    approval_instance.mayor_approval_date = timezone.now()
                    approval_instance.current_approval_level = 'completed'
                    registration.status = 'completed'
                    registration.current_step = 'certificate_generated'
                    messages.success(request, 'Mayor approval submitted. Registration approved! Certificate generated.')
                elif approval_instance.mayor_status == 'rejected':
                    approval_instance.current_approval_level = 'rejected'
                    registration.status = 'rejected'
                    messages.error(request, 'Mayor rejected the registration.')
                approval_instance.save()
                registration.save()
                if approval_instance.mayor_status == 'approved':
                    return redirect('land_management:generate_certificate', registration_id=registration.id)
                else:
                    return render(request, template_name, {
                        'form': form,
                        'registration': registration,
                        'approval': approval_instance,
                        'current_level': current_level,
                        'survey_payment': survey_payment,
                        'land_survey': land_survey,
                        'tax_payment': tax_payment,
                        'land_mapping': land_mapping,
                    })
            else:
                messages.error(request, 'An unexpected error occurred during approval state transition.')
                return render(request, template_name, {
                    'form': form,
                    'registration': registration,
                    'approval': approval_instance,
                    'current_level': current_level,
                    'survey_payment': survey_payment,
                    'land_survey': land_survey,
                    'tax_payment': tax_payment,
                    'land_mapping': land_mapping,
                })
        else:
            messages.error(request, 'Please correct the errors in the form.')
            return render(request, template_name, {
                'form': form,
                'registration': registration,
                'approval': approval_instance,
                'current_level': current_level,
                'survey_payment': survey_payment,
                'land_survey': land_survey,
                'tax_payment': tax_payment,
                'land_mapping': land_mapping,
            })
    else: # GET request
        user_full_name = request.user.get_full_name() or request.user.username
        user_email = request.user.email
        if current_level == 'director_pending':
            initial = {
                'director_full_name': user_full_name,
                'director_email': user_email,
                'director_title': 'Director',
            }
            form = DirectorApprovalForm(instance=approval_instance, initial=initial)
        elif current_level == 'secretary_pending' and approval_instance.director_status == 'approved':
            initial = {
                'secretary_full_name': user_full_name,
                'secretary_email': user_email,
                'secretary_title': 'Secretary',
            }
            form = SecretaryApprovalForm(instance=approval_instance, initial=initial)
        elif current_level == 'deputy_mayor_pending' and approval_instance.secretary_status == 'approved':
            initial = {
                'deputy_mayor_full_name': user_full_name,
                'deputy_mayor_email': user_email,
                'deputy_mayor_title': 'Deputy Mayor',
            }
            form = DeputyMayorApprovalForm(instance=approval_instance, initial=initial)
        elif current_level == 'mayor_pending' and approval_instance.deputy_mayor_status == 'approved':
            initial = {
                'mayor_full_name': user_full_name,
                'mayor_email': user_email,
                'mayor_title': 'Mayor',
            }
            form = MayorApprovalForm(instance=approval_instance, initial=initial)
        else:
            messages.info(request, f'Approval process is at: {current_level.replace("_", " ").title()}')
            return render(request, template_name, {
                'registration': registration,
                'approval': approval_instance,
                'current_level': current_level,
                'form': None,
                'survey_payment': survey_payment,
                'land_survey': land_survey,
                'tax_payment': tax_payment,
                'land_mapping': land_mapping,
            })

    return render(request, template_name, {
        'form': form,
        'registration': registration,
        'approval': approval_instance,
        'current_level': current_level,
        'survey_payment': survey_payment,
        'land_survey': land_survey,
        'tax_payment': tax_payment,
        'land_mapping': land_mapping,
    })

@login_required
def list_survey_payments(request):
    # Get all survey payments for the current user's registrations
    survey_payments = SurveyPayment.objects.filter(land_registration__user=request.user).select_related('land_registration')
    
    # Get all registrations that don't have survey payments yet
    registrations_without_payment = LandRegistration.objects.filter(
        user=request.user
    ).exclude(
        surveypayment__isnull=False
    )
    
    context = {
        'survey_payments': survey_payments,
        'registrations_without_payment': registrations_without_payment,
    }
    return render(request, 'land_management/payments/survey_payment_list.html', context)

@login_required
def list_land_surveys(request):
    # Get all land surveys for the current user's registrations
    land_surveys = LandSurvey.objects.filter(land_registration__user=request.user)
    
    # Get all registrations that have a survey payment but no land survey yet
    registrations_without_survey = LandRegistration.objects.filter(
        user=request.user,
        surveypayment__isnull=False
    ).exclude(
        landsurvey__isnull=False
    )

    context = {
        'land_surveys': land_surveys,
        'registrations_without_survey': registrations_without_survey,
    }
    return render(request, 'land_management/land_details/land_survey_list.html', context)

@login_required
def list_tax_payments(request):
    # Get all tax payments for the current user's registrations
    tax_payments = TaxPayment.objects.filter(land_registration__user=request.user)
    
    # Get all registrations that have a land survey but no tax payment yet
    registrations_without_tax_payment = LandRegistration.objects.filter(
        user=request.user,
        landsurvey__isnull=False
    ).exclude(
        taxpayment__isnull=False
    )

    context = {
        'tax_payments': tax_payments,
        'registrations_without_tax_payment': registrations_without_tax_payment,
    }
    return render(request, 'land_management/payments/tax_payment_list.html', context)

@login_required
def list_land_mappings(request):
    # Get all land mappings for the current user's registrations
    land_mappings = LandMapping.objects.filter(land_registration__user=request.user)
    
    # Get all registrations that have a tax payment but no land mapping yet
    registrations_without_mapping = LandRegistration.objects.filter(
        user=request.user,
        taxpayment__isnull=False
    ).exclude(
        landmapping__isnull=False
    )

    context = {
        'land_mappings': land_mappings,
        'registrations_without_mapping': registrations_without_mapping,
    }
    return render(request, 'land_management/land_details/land_mapping_list.html', context)

@login_required
def list_director_approvals(request):
    # Registrations truly awaiting *initial* director approval OR returned to director level
    pending_approvals_registrations = LandRegistration.objects.filter(
        current_step='land_mapping', 
        status='pending'
    ).filter(
        Q(approval__isnull=True) | Q(approval__director_status='pending') | Q(approval__director_status='returned')
    )

    # Approvals that the Director has already acted upon (approved, rejected, or returned)
    existing_director_actions = Approval.objects.filter(
        Q(director_status='approved') | Q(director_status='rejected') | Q(director_status='returned')
    ).select_related('land_registration')

    context = {
        'pending_approvals_registrations': pending_approvals_registrations,
        'existing_director_actions': existing_director_actions, # Renaming for clarity in template
    }
    return render(request, 'land_management/approvals/director_approval_list.html', context)

@login_required
def list_secretary_approvals(request):
    # Get registrations that have director approval and are awaiting secretary approval
    pending_approvals_registrations = LandRegistration.objects.filter(
        approval__director_status='approved',
        approval__current_approval_level='secretary_pending',
        approval__secretary_status='pending'
    )

    # Existing approval records where the secretary has taken an action
    existing_secretary_actions = Approval.objects.filter(
        Q(secretary_status='approved') | Q(secretary_status='rejected') | Q(secretary_status='returned')
    ).select_related('land_registration')

    context = {
        'pending_approvals_registrations': pending_approvals_registrations,
        'existing_secretary_actions': existing_secretary_actions,
    }
    return render(request, 'land_management/approvals/secretary_approval_list.html', context)

@login_required
def list_deputy_mayor_approvals(request):
    pending_approvals_registrations = LandRegistration.objects.filter(
        approval__secretary_status='approved',
        approval__current_approval_level='deputy_mayor_pending',
        approval__deputy_mayor_status='pending'
    )

    existing_deputy_mayor_actions = Approval.objects.filter(
        Q(deputy_mayor_status='approved') | Q(deputy_mayor_status='rejected') | Q(deputy_mayor_status='returned')
    ).select_related('land_registration')

    context = {
        'pending_approvals_registrations': pending_approvals_registrations,
        'existing_deputy_mayor_actions': existing_deputy_mayor_actions,
    }
    return render(request, 'land_management/approvals/deputy_mayor_approval_list.html', context)

@login_required
def list_mayor_approvals(request):
    pending_approvals_registrations = LandRegistration.objects.filter(
        approval__deputy_mayor_status='approved',
        approval__current_approval_level='mayor_pending',
        approval__mayor_status='pending'
    )

    existing_mayor_actions = Approval.objects.filter(
        Q(mayor_status='approved') | Q(mayor_status='rejected') | Q(mayor_status='returned')
    ).select_related('land_registration')

    context = {
        'pending_approvals_registrations': pending_approvals_registrations,
        'existing_mayor_actions': existing_mayor_actions,
    }
    return render(request, 'land_management/approvals/mayor_approvals_list.html', context)

@login_required
def registration_detail(request, registration_id):
    registration = get_object_or_404(LandRegistration, id=registration_id, user=request.user)
    
    # Fetch all related data
    survey_payment = getattr(registration, 'surveypayment', None)
    land_survey = getattr(registration, 'landsurvey', None)
    tax_payment = getattr(registration, 'taxpayment', None)
    land_mapping = getattr(registration, 'landmapping', None)
    approval = getattr(registration, 'approval', None)

    context = {
        'registration': registration,
        'survey_payment': survey_payment,
        'land_survey': land_survey,
        'tax_payment': tax_payment,
        'land_mapping': land_mapping,
        'approval': approval,
    }
    return render(request, 'land_management/registrations/registration_detail.html', context)

@login_required
def certificate_list(request):
    # Get all registrations that have completed the approval process
    registrations = LandRegistration.objects.filter(
        approval__mayor_status='approved'
    ).select_related(
        'approval',
        'surveypayment',
        'landsurvey',
        'taxpayment',
        'landmapping'
    ).order_by('-approval__mayor_approval_date')

    context = {
        'registrations': registrations,
    }
    return render(request, 'land_management/certificates/certificate_list.html', context)

@login_required
def generate_certificate(request, registration_id):
    registration = get_object_or_404(LandRegistration, id=registration_id)

    # Check if all steps are completed and final approval is granted
    if not all([
        registration.surveypayment,
        registration.landsurvey,
        registration.taxpayment,
        registration.landmapping,
        registration.approval,
        registration.approval.mayor_status == 'approved'
    ]):
        messages.error(request, 'Cannot generate certificate: Not all steps are completed or final approval is not granted.')
        return redirect('land_management:certificate_list')
    
    # Update registration status and current step to reflect certificate generation
    registration.status = 'completed'
    registration.current_step = 'certificate_generated'
    registration.save()

    context = {
        'registration': registration,
        'survey_payment': registration.surveypayment,
        'land_survey': registration.landsurvey,
        'tax_payment': registration.taxpayment,
        'land_mapping': registration.landmapping,
        'approval': registration.approval,
    }
    return render(request, 'land_management/certificates/certificate.html', context)

@login_required
def download_certificate_pdf(request, registration_id):
    registration = get_object_or_404(LandRegistration, id=registration_id)
    if not all([
        registration.surveypayment,
        registration.landsurvey,
        registration.taxpayment,
        registration.landmapping,
        registration.approval,
        registration.approval.mayor_status == 'approved'
    ]):
        messages.error(request, 'Cannot print certificate: Not all steps are completed or final approval is not granted.')
        return redirect('land_management:certificate_list')

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    elements.append(Paragraph('LAND REGISTRATION CERTIFICATE', styles['Title']))
    elements.append(Paragraph(f'Certificate No: {registration.transaction_reference}', styles['Normal']))
    elements.append(Paragraph(f'Issued on: {timezone.now().strftime("%Y-%m-%d")}', styles['Normal']))
    elements.append(Spacer(1, 12))

    # Gather all details into a single table
    details = [
        ['ID', registration.id],
        ['Transaction Reference', registration.transaction_reference],
        ['Land Code', registration.land_code],
        ['Owner Name', registration.buyer_full_name],
        ['Size', f'{registration.land_size} {registration.size_unit}'],
        ['Location', f'{registration.land_zone}, {registration.land_district}, {registration.land_region}'],
        ['Sale Price', registration.sale_price],
        ['Registration Date', registration.register_date.strftime('%Y-%m-%d') if registration.register_date else ''],
        ['Admin Name', registration.surveypayment.admin_name],
        ['Payer Name', registration.surveypayment.payer_name],
        ['Survey Payment Amount', registration.surveypayment.payment_amount],
        ['Survey Payment Method', registration.surveypayment.payment_method],
        ['Survey Payment Date', registration.surveypayment.payment_date.strftime('%Y-%m-%d') if registration.surveypayment.payment_date else ''],
        ['Survey Payment Status', registration.surveypayment.payment_status],
        ['Survey Payment Receipt', registration.surveypayment.payment_receipt.url if registration.surveypayment.payment_receipt else 'N/A'],
        ['Survey Number', registration.landsurvey.survey_number],
        ['Parcel Number', registration.landsurvey.parcel_number],
        ['Survey Land Code', registration.landsurvey.land_code],
        ['Survey Owner Name', registration.landsurvey.owner_name],
        ['Survey Date', registration.landsurvey.survey_date.strftime('%Y-%m-%d') if registration.landsurvey.survey_date else ''],
        ['Surveyor Name', registration.landsurvey.surveyor_name],
        ['Survey Location', registration.landsurvey.survey_location],
        ['Coordinates', registration.landsurvey.coordinates],
        ['Land Direction', registration.landsurvey.land_direction],
        ['Survey Documents', registration.landsurvey.survey_documents.url if registration.landsurvey.survey_documents else 'N/A'],
        ['Tax Admin Full Name', registration.taxpayment.admin_fullname],
        ['Tax Land Owner Name', registration.taxpayment.land_owner_name],
        ['Tax Land Reference No', registration.taxpayment.land_reference_no],
        ['Tax Land Price', registration.taxpayment.land_price],
        ['Tax Price', registration.taxpayment.tax_price],
        ['Tax Payment Date', registration.taxpayment.payment_date.strftime('%Y-%m-%d') if registration.taxpayment.payment_date else ''],
        ['Tax Payment Status', registration.taxpayment.payment_status],
        ['Tax Receipt Number', registration.taxpayment.receipt_number],
        ['Tax Notes', registration.taxpayment.notes],
        ['Land Mapping Reference', registration.landmapping.land_reference],
        ['Map Coordinates', registration.landmapping.map_coordinates],
        ['Mapping Date', registration.landmapping.mapping_date.strftime('%Y-%m-%d') if registration.landmapping.mapping_date else ''],
        ['Mapped By', registration.landmapping.mapped_by],
        ['Mapping Status', registration.landmapping.mapping_status],
        ['Map Document', registration.landmapping.map_document.url if registration.landmapping.map_document else 'N/A'],
        ['Director', f"{registration.approval.director_full_name} (Title: {registration.approval.director_title}, Email: {registration.approval.director_email})"],
        ['Director Status', f"{registration.approval.director_status} on {registration.approval.director_approval_date.strftime('%Y-%m-%d') if registration.approval.director_approval_date else ''}"],
        ['Secretary', f"{registration.approval.secretary_full_name} (Title: {registration.approval.secretary_title}, Email: {registration.approval.secretary_email})"],
        ['Secretary Status', f"{registration.approval.secretary_status} on {registration.approval.secretary_approval_date.strftime('%Y-%m-%d') if registration.approval.secretary_approval_date else ''}"],
        ['Deputy Mayor', f"{registration.approval.deputy_mayor_full_name} (Title: {registration.approval.deputy_mayor_title}, Email: {registration.approval.deputy_mayor_email})"],
        ['Deputy Mayor Status', f"{registration.approval.deputy_mayor_status} on {registration.approval.deputy_mayor_approval_date.strftime('%Y-%m-%d') if registration.approval.deputy_mayor_approval_date else ''}"],
        ['Mayor', f"{registration.approval.mayor_full_name} (Title: {registration.approval.mayor_title}, Email: {registration.approval.mayor_email})"],
        ['Mayor Status', f"{registration.approval.mayor_status} on {registration.approval.mayor_approval_date.strftime('%Y-%m-%d') if registration.approval.mayor_approval_date else ''}"],
    ]
    details_table = RLTable(details, hAlign='LEFT', colWidths=[160, 320])
    details_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(details_table)
    elements.append(Spacer(1, 18))
    elements.append(Paragraph('This certificate confirms the completion of the land registration process.', styles['Italic']))
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="certificate_{registration.transaction_reference}.pdf"'
    return response

@login_required
def report(request):
    from django.db.models import Q
    registrations = LandRegistration.objects.all().select_related('user')
    # Filtering
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status = request.GET.get('status')
    region = request.GET.get('region')
    if start_date:
        registrations = registrations.filter(register_date__gte=start_date)
    if end_date:
        registrations = registrations.filter(register_date__lte=end_date)
    if status and status != 'all':
        registrations = registrations.filter(status=status)
    if region and region != 'all':
        registrations = registrations.filter(land_region__icontains=region)

    # Summary stats - Define these before Excel export
    total = registrations.count()
    completed = registrations.filter(status='completed').count()
    pending = registrations.filter(status='pending').count()
    approved = registrations.filter(status='approved').count()
    rejected = registrations.filter(status='rejected').count()

    # Calculate total and average values
    total_value = registrations.aggregate(total=Sum('sale_price'))['total'] or 0
    avg_value = registrations.aggregate(avg=Avg('sale_price'))['avg'] or 0

    # Get region counts for chart
    region_counts = registrations.values('land_region').annotate(count=Count('id')).order_by('-count')

    all_regions = LandRegistration.objects.values_list('land_region', flat=True).distinct()

    # Export to Excel (enhanced: summary, charts, conditional formatting, flags, filters)
    if request.GET.get('export') == 'excel':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo
            from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)

            # 1. Summary Sheet
            summary = wb.create_sheet('Summary & Charts')
            summary['A1'] = 'LAND MANAGEMENT SYSTEM - REPORT SUMMARY'
            summary['A1'].font = Font(size=16, bold=True)
            summary['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            # Key metrics
            total = registrations.count()
            completed = registrations.filter(status='completed').count()
            pending = registrations.filter(status='pending').count()
            approved = registrations.filter(status='approved').count()
            rejected = registrations.filter(status='rejected').count()
            total_income = sum([
                (r.surveypayment.payment_amount if hasattr(r, 'surveypayment') and r.surveypayment and r.surveypayment.payment_amount else 0) +
                (r.taxpayment.tax_price if hasattr(r, 'taxpayment') and r.taxpayment and r.taxpayment.tax_price else 0)
                for r in registrations
            ])
            summary['A5'] = 'Total Registrations:'
            summary['B5'] = total
            summary['A6'] = 'Completed:'
            summary['B6'] = completed
            summary['A7'] = 'Pending:'
            summary['B7'] = pending
            summary['A8'] = 'Approved:'
            summary['B8'] = approved
            summary['A9'] = 'Rejected:'
            summary['B9'] = rejected
            summary['A10'] = 'Total Income:'
            summary['B10'] = total_income
            # Status distribution chart data
            chart_data = [
                ['Status', 'Count'],
                ['Completed', completed],
                ['Pending', pending],
                ['Approved', approved],
                ['Rejected', rejected],
            ]
            for i, row in enumerate(chart_data, 12):
                summary.append(row) if i == 12 else summary.append(row)
            # (Optional: Add chart code here if you want to use openpyxl's charting)

            # 2. Data Sheet
            ws = wb.create_sheet('Detailed Data')
            headers = [
                'Transaction Ref', 'Buyer', 'Seller', 'Land Code', 'Status', 'Region', 'Register Date',
                'Survey Payment Date', 'Survey Payment Amount',
                'Tax Payment Date', 'Tax Payment Amount',
                'Total Income',
                'Land Survey Date', 'Land Mapping Date',
                'Director Approval Date', 'Secretary Approval Date', 'Deputy Mayor Approval Date', 'Mayor Approval Date',
                'Total Days', 'Efficiency (%)', 'Attention Needed'
            ]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            # Data rows
            row_idx = 2
            for reg in registrations:
                survey_payment = getattr(reg, 'surveypayment', None)
                tax_payment = getattr(reg, 'taxpayment', None)
                land_survey = getattr(reg, 'landsurvey', None)
                land_mapping = getattr(reg, 'landmapping', None)
                approval = getattr(reg, 'approval', None)
                reg_date = reg.register_date.strftime('%Y-%m-%d') if reg.register_date else ''
                survey_payment_date = survey_payment.payment_date.strftime('%Y-%m-%d') if survey_payment and survey_payment.payment_date else ''
                tax_payment_date = tax_payment.payment_date.strftime('%Y-%m-%d') if tax_payment and tax_payment.payment_date else ''
                land_survey_date = land_survey.survey_date.strftime('%Y-%m-%d') if land_survey and land_survey.survey_date else ''
                land_mapping_date = land_mapping.mapping_date.strftime('%Y-%m-%d') if land_mapping and land_mapping.mapping_date else ''
                director_approval_date = approval.director_approval_date.strftime('%Y-%m-%d') if approval and approval.director_approval_date else ''
                secretary_approval_date = approval.secretary_approval_date.strftime('%Y-%m-%d') if approval and approval.secretary_approval_date else ''
                deputy_mayor_approval_date = approval.deputy_mayor_approval_date.strftime('%Y-%m-%d') if approval and approval.deputy_mayor_approval_date else ''
                mayor_approval_date = approval.mayor_approval_date.strftime('%Y-%m-%d') if approval and approval.mayor_approval_date else ''
                survey_payment_amount = float(survey_payment.payment_amount) if survey_payment and survey_payment.payment_amount else 0
                tax_payment_amount = float(tax_payment.tax_price) if tax_payment and tax_payment.tax_price else 0
                total_income = survey_payment_amount + tax_payment_amount
                total_days = ''
                if reg.register_date and approval and approval.mayor_approval_date:
                    total_days = (approval.mayor_approval_date - reg.register_date).days
                efficiency = ''
                if total_days != '' and total_days > 0:
                    expected_days = 30
                    efficiency = min(100, round((expected_days / total_days) * 100, 1))
                # Attention Needed: flag if pending and more than 30 days
                attention = ''
                if reg.status == 'pending' and reg.register_date:
                    days_since = (datetime.now().date() - reg.register_date).days
                    if days_since > 30:
                        attention = 'YES'
                ws.cell(row=row_idx, column=1, value=reg.transaction_reference)
                ws.cell(row=row_idx, column=2, value=reg.buyer_full_name)
                ws.cell(row=row_idx, column=3, value=reg.seller_full_name)
                ws.cell(row=row_idx, column=4, value=reg.land_code)
                ws.cell(row=row_idx, column=5, value=reg.status)
                ws.cell(row=row_idx, column=6, value=reg.land_region)
                ws.cell(row=row_idx, column=7, value=reg_date)
                ws.cell(row=row_idx, column=8, value=survey_payment_date)
                ws.cell(row=row_idx, column=9, value=survey_payment_amount)
                ws.cell(row=row_idx, column=10, value=tax_payment_date)
                ws.cell(row=row_idx, column=11, value=tax_payment_amount)
                ws.cell(row=row_idx, column=12, value=total_income)
                ws.cell(row=row_idx, column=13, value=land_survey_date)
                ws.cell(row=row_idx, column=14, value=land_mapping_date)
                ws.cell(row=row_idx, column=15, value=director_approval_date)
                ws.cell(row=row_idx, column=16, value=secretary_approval_date)
                ws.cell(row=row_idx, column=17, value=deputy_mayor_approval_date)
                ws.cell(row=row_idx, column=18, value=mayor_approval_date)
                ws.cell(row=row_idx, column=19, value=total_days)
                ws.cell(row=row_idx, column=20, value=efficiency)
                ws.cell(row=row_idx, column=21, value=attention)
                row_idx += 1
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            # Add filter to all columns
            ws.auto_filter.ref = ws.dimensions
            # Conditional formatting: highlight low efficiency and attention needed
            ws.conditional_formatting.add(f'T2:T{row_idx}', ColorScaleRule(start_type='min', start_color='FFEE1111', end_type='max', end_color='FF11EE11'))
            ws.conditional_formatting.add(f'U2:U{row_idx}', CellIsRule(operator='equal', formula=['"YES"'], fill=PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')))
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="registrations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"'
            wb.save(response)
            return response
        except ImportError:
            messages.error(request, 'Excel export requires openpyxl. Please install it.')
            return redirect('land_management:report')
    
    # Fallback CSV export function
    def export_csv(request, registrations, total, completed, pending, approved, rejected, total_value, avg_value, region_counts, all_regions, status, region, start_date, end_date):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="land_management_report_{}.csv"'.format(datetime.now().strftime('%Y%m%d_%H%M%S'))
        response.write(u'\ufeff'.encode('utf8'))
        writer = csv.writer(response)
        
        writer.writerow(['LAND MANAGEMENT SYSTEM - COMPREHENSIVE REPORT'])
        writer.writerow(['Generated on:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        writer.writerow([''])
        writer.writerow(['SUMMARY STATISTICS'])
        writer.writerow(['Total Registrations', total])
        writer.writerow(['Completed', completed, f"{completed/total*100:.1f}%" if total > 0 else "0%"])
        writer.writerow(['Pending', pending, f"{pending/total*100:.1f}%" if total > 0 else "0%"])
        writer.writerow(['Approved', approved, f"{approved/total*100:.1f}%" if total > 0 else "0%"])
        writer.writerow(['Rejected', rejected, f"{rejected/total*100:.1f}%" if total > 0 else "0%"])
        writer.writerow(['Total Revenue (SLS)', f"{total_value:,.2f}"])
        writer.writerow(['Average Sale Price (SLS)', f"{avg_value:,.2f}"])
        writer.writerow([''])
        writer.writerow(['REGION BREAKDOWN'])
        for region_data in region_counts:
            writer.writerow([region_data['land_region'], region_data['count']])
        writer.writerow([''])
        writer.writerow(['DETAILED REGISTRATION DATA'])
        
        # Enhanced headers
        writer.writerow([
            'Transaction Reference', 'Registration Date', 'Buyer Name', 'Seller Name',
            'Land Code', 'Land Region', 'Sale Price (SLS)', 'Status', 'Current Step',
            'Survey Payment Status', 'Tax Payment Status', 'Processing Days', 'Efficiency (%)'
        ])
        
        for reg in registrations:
            survey_payment = getattr(reg, 'surveypayment', None)
            tax_payment = getattr(reg, 'taxpayment', None)
            
            # Calculate processing days
            processing_days = 0
            if reg.register_date:
                processing_days = (datetime.now().date() - reg.register_date).days
            
            # Calculate efficiency
            efficiency = 0
            if reg.status == 'completed':
                efficiency = 100
            elif reg.status == 'approved':
                efficiency = 75
            elif reg.status == 'pending':
                if getattr(reg, 'landmapping', None):
                    efficiency = 60
                elif getattr(reg, 'landsurvey', None):
                    efficiency = 40
                elif tax_payment:
                    efficiency = 30
                elif survey_payment:
                    efficiency = 20
                else:
                    efficiency = 10
            
            writer.writerow([
                reg.transaction_reference,
                reg.register_date.strftime('%Y-%m-%d') if reg.register_date else '',
                reg.buyer_full_name,
                reg.seller_full_name,
                reg.land_code,
                reg.land_region,
                "{:.2f}".format(reg.sale_price) if reg.sale_price is not None else '',
                reg.status.title(),
                reg.current_step,
                survey_payment.payment_status if survey_payment else 'Not Started',
                tax_payment.payment_status if tax_payment else 'Not Started',
                processing_days,
                "{:.1f}".format(efficiency)
            ])
        return response

    if request.GET.get('export') == 'pdf':
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph('LAND MANAGEMENT SYSTEM - REPORT SUMMARY', styles['Title']))
        elements.append(Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', styles['Normal']))
        elements.append(Spacer(1, 12))
        # Summary
        summary_data = [
            ['Total Registrations', total],
            ['Completed', completed],
            ['Pending', pending],
            ['Approved', approved],
            ['Rejected', rejected],
            ['Total Value (SLS)', f"{total_value:,.2f}"],
            ['Average Sale Price (SLS)', f"{avg_value:,.2f}"],
        ]
        summary_table = RLTable(summary_data, hAlign='LEFT')
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 18))
        # Table headers
        data = [[
            'Transaction Ref', 'Buyer', 'Seller', 'Land Code', 'Status', 'Region', 'Register Date', 'Sale Price (SLS)'
        ]]
        for reg in registrations:
            data.append([
                reg.transaction_reference,
                reg.buyer_full_name,
                reg.seller_full_name,
                reg.land_code,
                reg.status,
                reg.land_region,
                reg.register_date.strftime('%Y-%m-%d') if reg.register_date else '',
                f"{reg.sale_price:,.2f}" if reg.sale_price is not None else '',
            ])
        table = RLTable(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="land_management_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"'
        return response

    context = {
        'registrations': registrations,
        'total': total,
        'completed': completed,
        'pending': pending,
        'approved': approved,
        'rejected': rejected,
        'total_value': total_value,
        'avg_value': avg_value,
        'region_counts': region_counts,
        'all_regions': all_regions,
        'selected_status': status or 'all',
        'selected_region': region or 'all',
        'start_date': start_date or '',
        'end_date': end_date or '',
    }
    return render(request, 'land_management/general/report.html', context)

@user_passes_test(lambda u: u.is_superuser)
@login_required
def profile(request):
    if request.method == 'POST':
        form = ProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('land_management:profile')
    else:
        form = ProfileForm(instance=request.user)
    return render(request, 'land_management/general/profile.html', {'form': form})

@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important!
            messages.success(request, 'Your password was successfully updated!')
            return redirect('land_management:profile')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'land_management/general/change_password.html', {'form': form})

@staff_member_required
def user_list(request):
    users = User.objects.all()
    return render(request, 'land_management/general/user_list.html', {'users': users})

@staff_member_required
def user_edit(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'User updated successfully!')
            return redirect('land_management:user_list')
    else:
        form = UserEditForm(instance=user)
    return render(request, 'land_management/general/user_edit.html', {'form': form, 'user_obj': user})

@staff_member_required
def user_reset_password(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    if request.method == 'POST':
        # Send password reset email
        from django.contrib.auth.forms import PasswordResetForm
        reset_form = PasswordResetForm({'email': user.email})
        if reset_form.is_valid():
            reset_form.save(
                request=request,
                use_https=request.is_secure(),
                email_template_name='land_management/general/password_reset_email.html',
                subject_template_name='land_management/general/password_reset_subject.txt',
            )
            messages.success(request, f'Password reset email sent to {user.email}')
            return redirect('land_management:user_list')
        else:
            messages.error(request, 'Could not send reset email.')
    return render(request, 'land_management/general/user_reset_password.html', {'user_obj': user})

def password_reset_request(request):
    user_not_found = False
    pending_request = None
    approved_request = None
    set_password_form = None
    user = None
    if request.method == 'POST':
        form = UsernamePasswordResetForm(request.POST)
        username = request.POST.get('username')
        # Check if this is a set password form submission
        if 'new_password1' in request.POST and 'new_password2' in request.POST and username:
            try:
                user = User.objects.get(username=username)
                approved_request = PasswordResetRequest.objects.filter(user=user, status='approved').first()
                if approved_request:
                    set_password_form = SetPasswordForm(user, request.POST)
                    if set_password_form.is_valid():
                        set_password_form.save()
                        approved_request.status = 'completed'
                        approved_request.save()
                        messages.success(request, 'Your password has been reset successfully. You can now log in with your new password.')
                        return redirect('land_management:login')
                else:
                    messages.error(request, 'No approved password reset request found for this user.')
            except User.DoesNotExist:
                user_not_found = True
                messages.error(request, 'User does not exist. Please check your username and try again.')
            return render(request, 'land_management/general/password_reset_form.html', {'form': form, 'user_not_found': user_not_found, 'set_password_form': set_password_form, 'approved_request': approved_request})
        # Otherwise, handle the request form as before
        if form.is_valid():
            try:
                user = form.get_user()
            except User.DoesNotExist:
                user_not_found = True
                messages.error(request, 'User does not exist. Please check your username and try again.')
                return render(request, 'land_management/general/password_reset_form.html', {'form': form, 'user_not_found': user_not_found})
            existing_pending_request = PasswordResetRequest.get_pending_request(user)
            if existing_pending_request:
                time_since_request = timezone.now() - existing_pending_request.requested_at
                hours_since_request = time_since_request.total_seconds() / 3600
                if hours_since_request < 24:
                    messages.warning(
                        request,
                        f'You already have a pending password reset request submitted on {existing_pending_request.requested_at.strftime("%B %d, %Y at %I:%M %p")}. '
                        'Please wait for a superuser or administrator to review and approve your request. '
                        'You will receive an email notification once your request is processed. '
                        'If you need urgent assistance, please contact the system administrator directly.'
                    )
                    return redirect('land_management:login')
                else:
                    existing_pending_request.status = 'rejected'
                    existing_pending_request.rejection_reason = 'Request expired - user submitted new request'
                    existing_pending_request.processed_at = timezone.now()
                    existing_pending_request.save()
            reset_request = PasswordResetRequest.objects.create(
                user=user,
                rejection_reason=form.cleaned_data.get('reason', '')
            )
            superusers = User.objects.filter(is_superuser=True)
            for superuser in superusers:
                Notification.objects.create(
                    title=f"Password Reset Request from {user.username}",
                    message=f"User {user.get_full_name() or user.username} has requested a password reset. Reason: {form.cleaned_data.get('reason', 'No reason provided')}",
                    notification_type='password_reset',
                    user=superuser,
                    related_object_id=reset_request.id,
                    related_object_type='PasswordResetRequest'
                )
            messages.success(request, 'Your password reset request has been submitted and is pending admin approval. You will receive an email once it is approved.')
            return redirect('land_management:login')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = UsernamePasswordResetForm()
        username = request.GET.get('username')
        if username:
            try:
                user = User.objects.get(username=username)
                # Check for approved request
                approved_request = PasswordResetRequest.objects.filter(user=user, status='approved').first()
                if approved_request:
                    set_password_form = SetPasswordForm(user)
                else:
                    pending_request = PasswordResetRequest.get_pending_request(user)
            except User.DoesNotExist:
                user_not_found = True
    return render(request, 'land_management/general/password_reset_form.html', {
        'form': form,
        'pending_request': pending_request,
        'user_not_found': user_not_found,
        'set_password_form': set_password_form,
        'approved_request': approved_request,
        'username': username or '',
    })

@login_required
def registration_list(request):
    if request.user.is_staff:
        registrations = LandRegistration.objects.all()
    else:
        registrations = LandRegistration.objects.filter(user=request.user)
    return render(request, 'land_management/registrations/registration_list.html', {'registrations': registrations})

@login_required
def edit_land_registration(request, registration_id):
    registration = LandRegistration.objects.get(pk=registration_id)
    if not (request.user.is_staff or registration.user == request.user):
        return redirect('land_management:registration_list')
    if request.method == 'POST':
        form = LandRegistrationForm(request.POST, request.FILES, instance=registration)
        if form.is_valid():
            form.save()
            return redirect('land_management:registration_detail', registration_id=registration.id)
    else:
        form = LandRegistrationForm(instance=registration)
    return render(request, 'land_management/registrations/land_registration.html', {'form': form, 'edit_mode': True, 'registration': registration})

@login_required
def delete_land_registration(request, registration_id):
    registration = LandRegistration.objects.get(pk=registration_id)
    if not (request.user.is_staff or registration.user == request.user):
        return redirect('land_management:registration_list')
    if request.method == 'POST':
        registration.delete()
        return redirect('land_management:registration_list')
    return render(request, 'land_management/registrations/registration_confirm_delete.html', {'registration': registration})

@user_passes_test(lambda u: u.is_superuser)
def create_user(request):
    groups = Group.objects.all()
    if request.method == 'POST':
        form = UserCreateForm(request.POST)
        selected_group_id = request.POST.get('group')
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])
            user.save()
            # Assign to selected group if provided
            if selected_group_id:
                try:
                    group = Group.objects.get(id=selected_group_id)
                    user.groups.add(group)
                except Group.DoesNotExist:
                    pass
            messages.success(request, 'Admin user created successfully!')
            return redirect('land_management:user_list')
    else:
        form = UserCreateForm()
    return render(request, 'land_management/general/user_create.html', {'form': form, 'groups': groups})

@login_required
def gift_land_registration(request, registration_id=None):
    registration = None
    if registration_id:
        registration = get_object_or_404(LandRegistration, id=registration_id, user=request.user, registration_type='gift')

    if request.method == 'POST':
        try:
            if registration: # Editing an existing registration
                form = GiftLandRegistrationForm(request.POST, request.FILES, instance=registration)
            else: # Creating a new registration
                form = GiftLandRegistrationForm(request.POST, request.FILES)
            
            if form.is_valid():
                registration = form.save(commit=False)
                registration.user = request.user
                registration.registration_type = 'gift'  # <-- force the type
                registration.status = 'pending'
                registration.current_step = 'registration'

                # Generate unique transaction_reference only for new registrations
                if not registration_id:
                    registration.transaction_reference = f"GIFT-{timezone.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"

                registration.save()

                # If this was a returned registration, clear approval return info
                if registration_id and hasattr(registration, 'approval') and registration.approval.return_step:
                    approval_instance = registration.approval
                    approval_instance.return_step = None
                    approval_instance.rejection_reason = None
                    approval_instance.rejection_date = None
                    approval_instance.returned_by = None
                    approval_instance.save()

                messages.success(request, f'Gift land registration submitted successfully! Transaction Reference: {registration.transaction_reference}')
                return redirect('land_management:survey_payment', registration_id=registration.id)
            else:
                messages.error(request, 'Please correct the errors in the form.')
        except Exception as e:
            messages.error(request, f'Error saving registration: {str(e)}')
    else: # GET request
        if registration: # Pre-fill form for existing registration
            form = GiftLandRegistrationForm(instance=registration, initial={'registration_type': 'gift'})
        else: # Blank form for new registration
            form = GiftLandRegistrationForm(initial={'registration_type': 'gift'})

    return render(request, 'land_management/registrations/gift_land_registration.html', {'form': form, 'registration': registration})

@login_required
def inheritance_land_registration(request, registration_id=None):
    registration = None
    if registration_id:
        registration = get_object_or_404(LandRegistration, id=registration_id, user=request.user, registration_type='inheritance')

    if request.method == 'POST':
        try:
            if registration: # Editing an existing registration
                form = InheritanceLandRegistrationForm(request.POST, request.FILES, instance=registration)
            else: # Creating a new registration
                form = InheritanceLandRegistrationForm(request.POST, request.FILES)
            
            if form.is_valid():
                registration = form.save(commit=False)
                registration.user = request.user
                registration.registration_type = 'inheritance'  # <-- force the type
                registration.status = 'pending'
                registration.current_step = 'registration'

                # Generate unique transaction_reference only for new registrations
                if not registration_id:
                    registration.transaction_reference = f"INH-{timezone.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"

                registration.save()

                # If this was a returned registration, clear approval return info
                if registration_id and hasattr(registration, 'approval') and registration.approval.return_step:
                    approval_instance = registration.approval
                    approval_instance.return_step = None
                    approval_instance.rejection_reason = None
                    approval_instance.rejection_date = None
                    approval_instance.returned_by = None
                    approval_instance.save()

                messages.success(request, f'Inheritance land registration submitted successfully! Transaction Reference: {registration.transaction_reference}')
                return redirect('land_management:survey_payment', registration_id=registration.id)
            else:
                messages.error(request, 'Please correct the errors in the form.')
        except Exception as e:
            messages.error(request, f'Error saving registration: {str(e)}')
    else: # GET request
        if registration: # Pre-fill form for existing registration
            form = InheritanceLandRegistrationForm(instance=registration, initial={'registration_type': 'inheritance'})
        else: # Blank form for new registration
            form = InheritanceLandRegistrationForm(initial={'registration_type': 'inheritance'})

    return render(request, 'land_management/registrations/inheritance_land_registration.html', {'form': form, 'registration': registration})

@user_passes_test(lambda u: u.is_superuser)
@login_required
def password_reset_approval_dashboard(request):
    """Dashboard for superusers to approve/reject password reset requests"""
    pending_requests = PasswordResetRequest.objects.filter(status='pending').order_by('-requested_at')
    approved_requests = PasswordResetRequest.objects.filter(status='approved').order_by('-requested_at')[:10]
    rejected_requests = PasswordResetRequest.objects.filter(status='rejected').order_by('-requested_at')[:10]
    
    context = {
        'pending_requests': pending_requests,
        'approved_requests': approved_requests,
        'rejected_requests': rejected_requests,
    }
    return render(request, 'land_management/general/password_reset_approval_dashboard.html', context)

@user_passes_test(lambda u: u.is_superuser)
@login_required
def approve_password_reset(request, request_id):
    """Approve a password reset request (no email sent)"""
    reset_request = get_object_or_404(PasswordResetRequest, id=request_id, status='pending')
    try:
        # Get the user from the reset request
        user = reset_request.user
        # Update request status (no email sent)
        reset_request.status = 'approved'
        reset_request.processed_by = request.user
        reset_request.processed_at = timezone.now()
        reset_request.save()
        messages.success(request, f'Password reset approved for {user.username}. (No email sent)')
    except Exception as e:
        messages.error(request, f'Error approving password reset: {str(e)}')
    return redirect('land_management:password_reset_approval_dashboard')

@user_passes_test(lambda u: u.is_superuser)
@login_required
def reject_password_reset(request, request_id):
    """Reject a password reset request"""
    reset_request = get_object_or_404(PasswordResetRequest, id=request_id, status='pending')
    
    reset_request.status = 'rejected'
    reset_request.processed_by = request.user
    reset_request.processed_at = timezone.now()
    reset_request.save()
    
    messages.success(request, f'Password reset request for {reset_request.user.username} has been rejected')
    return redirect('land_management:password_reset_approval_dashboard')

@user_passes_test(lambda u: u.is_superuser)
@login_required
def notifications_view(request):
    """View all notifications for superusers"""
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
    
    # Mark notifications as read when viewed
    unread_notifications = notifications.filter(is_read=False)
    unread_notifications.update(is_read=True)
    
    context = {
        'notifications': notifications,
    }
    return render(request, 'land_management/general/notifications.html', context)

@user_passes_test(lambda u: u.is_superuser)
@login_required
def mark_notification_read(request, notification_id):
    """Mark a notification as read"""
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    notification.is_read = True
    notification.save()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': 'success'})
    
    return redirect('land_management:notifications')
